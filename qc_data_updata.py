# This Python file uses the following encoding: utf-8
import sys
import xlsxwriter
import datetime
import time
from openpyxl import load_workbook, Workbook

reload(sys)
sys.setdefaultencoding('utf-8')

today = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d')
# 用当前时间减去一天的秒数 得到昨天的时间
yesterday = datetime.datetime.fromtimestamp(time.time() - 86400).strftime('%Y-%m-%d')
print today
print yesterday
t_work = load_workbook(today + ".xlsx")
y_work = load_workbook(yesterday + ".xlsx")
t_sheet = t_work.get_sheet_by_name('Sheet1')
y_sheet = y_work.get_sheet_by_name('Sheet1')
# 获取表格所有行和列，两者都是可迭代的
t_rows = t_sheet.rows
y_rows = y_sheet.rows

##保存差异的数据
wb_chage = Workbook()
ws_chage = wb_chage.get_active_sheet()
# 记录保存数据的行数
row_chage = 2
ws_chage.cell(row=1, column=1).value = '商品id'
ws_chage.cell(row=1, column=2).value = '昨日价格'
ws_chage.cell(row=1, column=3).value = '今日价格'

for t_row in t_rows:
    ##当前商品ID
    curr_goods_id = t_row[0].value
    # 当前商品属性ID
    curr_product_id = t_row[1].value
    # 当前商品属性对应的价格
    curr_product_price = t_row[4].value
    print "当前商品信息:" + str(curr_goods_id) + '--' + str(curr_product_id) + '--' + str(
        curr_product_price)
    # while index_y_rows< y_rows.
    y_rows = y_sheet.rows
    for y_row in y_rows:
        goods_id = y_row[0].value
        product_id = y_row[1].value
        product_price = y_row[4].value
        print "循环商品信息:" + str(goods_id) + '--' + str(product_id) + '--' + str(product_price)
        if curr_goods_id == goods_id:
            if curr_product_id == product_id:
                if curr_product_price != product_price:
                    ws_chage.cell(row=row_chage, column=1).value = str(curr_product_id)
                    ws_chage.cell(row=row_chage, column=2).value = str(curr_product_price)
                    ws_chage.cell(row=row_chage, column=3).value = str(product_price)
                    row_chage += 1
                    print curr_product_id + "is chage"
wb_chage.save(filename=today + '_price_chage.xlsx')
print "=================校对完成============================="
